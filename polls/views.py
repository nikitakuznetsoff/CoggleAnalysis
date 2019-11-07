import json
import requests

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from django.views import generic
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth

from modules import checks, coggle, readers, analysis, text_search
from .models import UserData, Task, Homework


@login_required(login_url='/accounts/login/')
def index(request):
    #coggle.coggle_user.authorization()
    userData = UserData.objects.get(user=request.user)
    context = {'userData': userData}
    return render(request, 'polls/index.html', context)


class HomeworksView(LoginRequiredMixin, generic.DetailView):
    model = Task
    template_name = 'polls/homeworks.html'


class TasksView(LoginRequiredMixin, generic.DetailView):
    model = Homework
    template_name = 'polls/task.html'


@login_required(login_url='/accounts/login/')
def homework_add(request, task_id):
    context = { "task_id": task_id }
    return render(request, 'polls/homeworks_add.html', context)


@login_required(login_url='/accounts/login/')
def homework_add_confirm(request, task_id):
    try:
        name = request.POST['name']
        link = request.POST['link']
        link = readers.link_to_id(link)
    except Exception:
        return render(request, 'polls/actions/add_error.html')
    else:
        userData = UserData.objects.get(user=request.user)
        task = userData.task_set.get(pk=task_id)
        task.homework_set.create(name=name, link=link)
        return HttpResponseRedirect(reverse('add_done'))


class HomeworkDelete(LoginRequiredMixin, generic.DetailView):
    model = Task
    template_name = 'polls/homework_delete.html'


@login_required(login_url='/accounts/login/')
def homework_delete_page(request, task_id, homework_id):
    task = Task.objects.get(pk=task_id)
    homework = task.homework_set.get(pk=homework_id)
    context = { 'task_id': task_id,
                'homework': homework }
    return render(request, 'polls/homework_delete_page.html', context)


@login_required(login_url='/accounts/login/')
def homework_delete_confirm(request, task_id, homework_id):
    task = Task.objects.get(pk=task_id)
    homework = task.homework_set.get(pk=homework_id)
    try:
        homework.delete()
    except:
        context = { 'title': "Удаление работы",
                    'text': "Произошла ошибка, работы не удалена"}
        return render(request, 'polls/actions/error.html', context)
    else:
        return HttpResponseRedirect(reverse('delete_done'))

########################

class AddView(LoginRequiredMixin, generic.ListView):
    model = UserData
    template_name = 'polls/actions/add.html'


@login_required(login_url='/accounts/login/')
def add_confirm(request):
    try:
        title_new = request.POST['title']
        about_new = request.POST['about']
        file = request.FILES['file_table']
        table_name = request.POST['table']

        true_work = request.POST['true_work']
        true_work_link = readers.link_to_id(true_work)

        keys = request.POST['keys']
        names_cell = request.POST['names']
        links_cell = request.POST['links']

        wb = load_workbook(filename=file)
    except Exception:
        return render(request, 'polls/actions/add_error.html')
    else:
        # Проверка наличия заданной таблицы в Exel файле
        if table_name != '':
            if not checks.check_correct_tablename(wb, table_name):
                return render(request, 'polls/actions/add_error.html')

        # Проверка формата заданных ячеек
        if not checks.check_correct_cellname(names_cell) or not checks.check_correct_cellname(links_cell):
            return render(request, 'polls/actions/add_error.html')

        try:
            # Проверка наличия токена для анализа
            if coggle.coggle_user.access_token == "":
                coggle.coggle_user.authorization()
                return render(request, 'polls/actions/add.html')
        except Exception:
            return render(request, 'polls/actions/add_error.html')

        # Создание объекта домашнего задания для текущего пользователя
        curr_data = UserData.objects.get(user=request.user)
        curr_task = curr_data.task_set.create(title=title_new, about=about_new)

        # Создание массива ключевых значений
        arr_keys = keys.split(",")

        # Загрузка нужной таблицы
        wb = load_workbook(file)
        if table_name != '':
            sheet = wb[table_name]
        else:
            sheet = wb[wb.sheetnames[0]]

        # Чтение информации из таблицы
        arr = readers.read_mindmap_ids(sheet, names_cell, links_cell)

        # Получение массива коэффициентов структурного сходства
        arr_with_coef = analysis.mindmap_analysis(arr[1], true_work_link)

        # Массив текстов каждой интеллект карты
        arr_text = analysis.take_text(arr[1])

        # Массив с вычисленными сходствами текстовых составляющих
        if len(arr_keys) > 0:
            sim_arr = text_search.initialization(arr_keys, arr_text)
        else:
            sim_arr = [0] * len(arr_keys)


        # Создание работ
        for i in range(0, len(arr[0])):
            # Проверка на считывание лишней информации
            if arr[0][i] is None:
                break
            if true_work_link == "":
                coef = round(arr_with_coef[1][i] * 100)
            else:
                coef = round(arr_with_coef[i] * 100)

            if len(sim_arr) > 0:
                sim = round(sim_arr[i] * 100)
            else:
                sim = -1

            curr_task.homework_set.create(name=arr[0][i], link=arr[1][i],
                                          similarity=coef, plagiarism=sim)

        return HttpResponseRedirect(reverse('add_done'))


@login_required(login_url='/accounts/login/')
def add_done(request):
    return render(request, 'polls/actions/add_done.html')

########################

@login_required(login_url='/accounts/login/')
def change(request):
    userData = UserData.objects.get(user=request.user)
    context = {'userData': userData}
    return render(request, 'polls/actions/change.html', context)


@login_required(login_url='/accounts/login/')
def change_task(request, task_id):
    task = Task.objects.get(pk=task_id)
    context = {'task': task,
               'task_id' : task_id}
    return render(request, 'polls/actions/change_form.html', context)


@login_required(login_url='/accounts/login/')
def change_task_confirm(request, task_id):
    task = Task.objects.get(pk=task_id)
    try:
        title_new = request.POST['title']
        about_new = request.POST['about']
    except:
        return render(request, 'polls/actions/change_error.html')
    else:
        task.title = title_new
        task.about = about_new
        task.save()
        return HttpResponseRedirect(reverse('change_done'))


@login_required(login_url='/accounts/login/')
def change_done(request):
    return render(request, 'polls/actions/change_done.html')

########################

@login_required(login_url='/accounts/login/')
def delete(request):
    userData = UserData.objects.get(user=request.user)
    context = {'userData': userData}
    return render(request, 'polls/actions/delete.html', context)


@login_required(login_url='/accounts/login/')
def delete_task(request, task_id):
    task = Task.objects.get(pk=task_id)
    context = {'task': task,
               'task_id' : task_id}
    return render(request, 'polls/actions/delete_form.html', context)


@login_required(login_url='/accounts/login/')
def delete_task_confirm(request, task_id):
    task = Task.objects.get(pk=task_id)
    try:
        task.delete()
    except:
        return render(request, 'polls/actions/change_error.html')
    else:
        return HttpResponseRedirect(reverse('delete_done'))


@login_required(login_url='/accounts/login/')
def delete_done(request):
    return render(request, 'polls/actions/delete_done.html')

########################


@login_required(login_url='/accounts/login/')
def coggle_auth(request):
    try:
        code = request.GET['code']
    except Exception:
        context = { 'title': "Ошибка токена",
                    'text': "Ключ авторизации не был получен"}
        return render(request, 'polls/actions/error.html', context)
    else:
        params = {"code": code, "grant_type": "authorization_code", "redirect_uri": coggle.redirect_uri}
        resp1 = requests.post(coggle.coggle_user.url_base + "token", auth=HTTPBasicAuth(coggle.client_id, coggle.client_secret), json=params)
        information_auth = json.loads(resp1.text)
        coggle.coggle_user.access_token = information_auth["access_token"]

        context = {'title': "Получение токена",
                   'text': "Ключ авторизации был успешно получен"}
    return render(request, 'polls/actions/done.html', context)
