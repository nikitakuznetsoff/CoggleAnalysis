from main_site.models import UserData, Task, coggle, miro, MindMap
from main_site.forms import TaskForm

from openpyxl import load_workbook
import json
from modules import checks, readers, algorithm

from django.shortcuts import render
from django.contrib.auth.decorators import login_required


@login_required(login_url='/accounts/login/')
def index(request):
    try:
        user_data = UserData.objects.get(user=request.user)
    except UserData.DoesNotExist:
        user_data = ''
    context = {'userData': user_data}
    return render(request, 'main_site/index.html', context)


@login_required(login_url='/accounts/login/')
def info_task_view(request, task_id):
    try:
        task = Task.objects.get(pk=task_id)
        keys = ''
        if len(task.keys) > 0:
            keys = json.loads(task.keys)
    except Task.DoesNotExist:
        task = ''
        keys = ''
    context = {'task': task, 'keys': keys, 'task_id': task_id}
    return render(request, 'main_site/actions/homeworks/info.html', context)


@login_required(login_url='/accounts/login/')
def task_add(request):
    user_data = UserData.objects.get(user=request.user)
    context = dict()
    context['coggle_key'] = user_data.coggle_key == "Undefined"
    context['miro_key'] = user_data.miro_key == "Undefined"
    return render(request, 'main_site/actions/add.html', context)


@login_required(login_url='/accounts/login/')
def task_add_confirm(request):
    if request.method == 'POST':
        task_form = TaskForm(request.POST, request.FILES)
        if task_form.is_valid():
            file_table = request.FILES['file_table']
            # Загрузка нужной таблицы
            workbook = load_workbook(filename=file_table, read_only=True)

            # Вся информация из формы
            data = task_form.cleaned_data

            # Проверка наличия заданной таблицы в Excel файле
            if data['table_name'] != '':
                if not checks.check_correct_tablename(workbook, data['table_name']):
                    context = {
                        'title': "Создание задания",
                        'text': "Заданная таблица отсутствует в файле "
                    }
                    return render(request, 'main_site/actions/error.html', context)

            # Проверка формата заданных ячеек
            if data['names_column'] != '' and not checks.column_name_is_correct(data['names_column']) \
                    or data['links_column'] != '' and not checks.column_name_is_correct(data['links_column']) \
                    or data['start_row'] != '' and not checks.column_name_is_correct(data['start_row']):
                context = {
                    'title': "Создание задания",
                    'text': "Ошибка в формате столбцов / ячеек"
                }
                return render(request, 'main_site/actions/error.html', context)

            # Получение нужной таблицы из файла
            if data['table_name'] != '':
                sheet = workbook[data['table_name']]
            else:
                sheet = workbook[workbook.sheetnames[0]]
            print("[SHEET] " + str(sheet))

            # Чтение информации из таблицы
            mindmaps_table_info = readers.read_mindmaps_info(
                sheet=sheet,
                names_column=data['names_column'],
                links_column=data['links_column'],
                start_row=data['start_row']
            )
            print("Readed Names: " + str(mindmaps_table_info['names']))
            print("Readed Links: " + str(mindmaps_table_info['mindmaps_id']))

            # UserData текущего пользователя
            user_data = UserData.objects.get(user=request.user)

            # Инициализация объекта сервиса
            if data['service'] == "Coggle":
                web_service_object = coggle.Coggle(
                    app_name=coggle.APP_NAME,
                    client_id=coggle.CLIENT_ID,
                    client_secret=coggle.CLIENT_SECRET,
                    redirect_uri=coggle.REDIRECT_URI,
                    access_token=user_data.coggle_key
                )
            elif data['service'] == "Miro":
                web_service_object = miro.Miro(
                    app_name=miro.APP_NAME,
                    client_id=miro.CLIENT_ID,
                    client_secret=miro.CLIENT_SECRET,
                    redirect_uri=miro.REDIRECT_URI,
                    access_token=user_data.miro_key
                )
            else:
                context = {
                    'title': "Создание задания",
                    'text': "Ошибка в выборе сервиса"
                }
                return render(request, 'main_site/actions/error.html', context)

            # Создание объекта MindMap для каждой работы и инициализация графа
            mindmaps = []
            for i in range(len(mindmaps_table_info['names'])):
                mindmap = MindMap(
                    name=mindmaps_table_info['names'][i],
                    id=mindmaps_table_info['mindmaps_id'][i],
                    service=str(web_service_object)
                )
                print("MINDMAP: " + mindmap.name + "; ID: " + mindmap.id)
                response = web_service_object.nodes(mindmap.id)
                mindmap.create_graph_view(response)
                mindmaps.append(mindmap)

            correct_mindmap = None
            if data['correct_work'] != '':
                correct_mindmap = MindMap(
                    name="correct work",
                    id=readers.link_to_id(data['correct_work'].strip()),
                    service=str(web_service_object)
                )
                response = web_service_object.nodes(correct_mindmap.id)
                correct_mindmap.create_graph_view(response)
            print("CORRECT MINDMAP ID: " + correct_mindmap.id)

            if correct_mindmap is not None:
                for mindmap in mindmaps:
                    if mindmap.id == correct_mindmap.id:
                        mindmap.similarity_score = 1
                    else:
                        mindmap.similarity_score = algorithm.max_common_substree_rooted(
                            correct_mindmap.graph, mindmap.graph
                    )

            # Создание массива ключевых значений
            key_values = data['text_keys'].split(",")
            for i in range(len(key_values)):
                key_values[i] = key_values[i].strip()

            # Оценка текстового содержания
            for mindmap in mindmaps:
                text = mindmap.get_text()
                mindmap.text = text
                text_score = 0
                for v in key_values:
                    if v in text:
                        text_score += 1
                text_score /= 1. * len(key_values)
                mindmap.text_score = text_score

            # sim_arr = []
            # # Массив с вычисленными сходствами текстовых составляющих
            # if len(arr_keys) > 0:
            #     sim_arr = text_search.initialization(arr_keys, arr_text)
            # else:
            #     sim_arr = [0] * len(arr_keys)

            # Создание объекта задания для текущего пользователя
            curr_task = user_data.task_set.create(
                title=data['title'],
                about=data['about'],
                keys=json.dumps(key_values)
            )
            print("[CREATING TASK] TITLE: " + data['title'] + "; ABOUT: " + data['about'])

            # Создание работ студентов
            for mindmap in mindmaps:
                metrics = mindmap.get_metrics()
                keys = json.dumps(mindmap.text_keys)
                curr_task.homework_set.create(
                    name=mindmap.name,
                    link=mindmap.id,
                    service=mindmap.service,
                    # Metrics
                    similarity_score=mindmap.similarity_score * 100,
                    text_score=mindmap.text_score,
                    text_keys=keys,
                    plagiarism=mindmap.plagiarism,
                    count_nodes=metrics['count_nodes'],
                    count_first_layer_branches=metrics['count_first_layer_branches'],
                    average_node_text=metrics['avg_node_text_len'],
                    max_height=metrics['max_height']
                )

            context = {
                'title': "Создание задания",
                'text': "Задание успешно добавлено"
            }
            return render(request, 'main_site/actions/done.html', context)
        else:
            return render(request, 'main_site/actions/add.html', {'form': task_form})

    else:
        task_form = TaskForm()
    return render(request, 'main_site/actions/add.html', {'form': task_form})


########################

@login_required(login_url='/accounts/login/')
def change(request):
    try:
        user_data = UserData.objects.get(user=request.user)
    except UserData.DoesNotExist:
        user_data = ''
    context = {'userData': user_data}
    return render(request, 'main_site/actions/change.html', context)


@login_required(login_url='/accounts/login/')
def change_task(request, task_id):
    task = Task.objects.get(pk=task_id)
    context = {'task': task,
               'task_id': task_id}
    return render(request, 'main_site/actions/change_confirm.html', context)


@login_required(login_url='/accounts/login/')
def change_task_confirm(request, task_id):
    task = Task.objects.get(pk=task_id)
    try:
        title_new = request.POST['title']
        about_new = request.POST['about']
    except:
        context = {
            'title': "Изменение описания",
            'text': "Произошла ошибка, изменения не сохранены"
        }
        return render(request, 'main_site/actions/error.html', context)
    else:
        task.title = title_new
        task.about = about_new
        task.save()
        context = {
            'title': "Изменение описания",
            'text': "Изменения успешно сохранены"
        }
        return render(request, 'main_site/actions/done.html', context)


########################

@login_required(login_url='/accounts/login/')
def delete(request):
    try:
        user_data = UserData.objects.get(user=request.user)
    except UserData.DoesNotExist:
        user_data = ''
    context = {'userData': user_data}
    return render(request, 'main_site/actions/delete.html', context)


@login_required(login_url='/accounts/login/')
def delete_task(request, task_id):
    task = Task.objects.get(pk=task_id)
    context = {
        'task': task,
        'task_id': task_id
    }
    return render(request, 'main_site/actions/delete_confirm.html', context)


@login_required(login_url='/accounts/login/')
def delete_task_confirm(request, task_id):
    task = Task.objects.get(pk=task_id)
    try:
        task.delete()
    except:
        context = {
            'title': "Удаление задание",
            'text': "Произошла ошибка при удалении задания"
        }
        return render(request, 'main_site/actions/error.html', context)
    else:
        context = {
            'title': "Удаление задания",
            'text': "Задание успешно удалено"
        }
        return render(request, 'main_site/actions/done.html', context)
