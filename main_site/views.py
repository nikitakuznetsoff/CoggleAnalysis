import json
import requests

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, redirect
from django.views import generic


from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth

from modules import checks, coggle, miro, readers, analysis, text_search
from .models import UserData, Task, Homework

from .forms import TaskForm


@login_required(login_url='/accounts/login/')
def index(request):
    # coggle.coggle_client.authorization()
    try:
        user_data = UserData.objects.get(user=request.user)
    except UserData.DoesNotExist:
        user_data = ''
    context = {'userData': user_data}
    return render(request, 'main_site/index.html', context)


@login_required(login_url='/accounts/login/')
def homeworks(request, task_id):
    try:
        task = Task.objects.get(pk=task_id)
    except Task.DoesNotExist:
        task = ''
    context = {
        'task_id': task_id,
        'task': task
    }
    return render(request, 'main_site/actions/homeworks/list.html', context)


@login_required(login_url='/accounts/login/')
def homework(request, task_id, homework_id):
    try:
        hw = Homework.objects.get(pk=homework_id)
    except Homework.DoesNotExist:
        hw = ''
    context = {
        'task_id': task_id,
        'homework': hw
    }
    return render(request, 'main_site/actions/homeworks/homework.html', context)


@login_required(login_url='/accounts/login/')
def homework_add(request, task_id):
    context = {
        'task_id': task_id
    }
    return render(request, 'main_site/actions/homeworks/create.html', context)


@login_required(login_url='/accounts/login/')
def homework_add_confirm(request, task_id):
    try:
        name = request.POST['name']
        link = request.POST['link']
        link = readers.link_to_id(link)
    except Exception as e:
        context = {
            'title': "Добавление работы",
            'text': "Произошла ошибка при добавлении работы"
        }
        return render(request, 'main_site/actions/error.html', context)
    else:
        user_data = UserData.objects.get(user=request.user)
        task = user_data.task_set.get(pk=task_id)
        task.homework_set.create(name=name, link=link)

        context = {
            'title': "Добавление работы",
            'text': "Работа успешно добалена"
        }
        return render(request, 'main_site/actions/done.html', context)


@login_required(login_url='/accounts/login/')
def homework_delete(request, task_id):
    task = Task.objects.get(pk=task_id)
    context = {
        'task_id': task_id,
        'task': task,
    }
    return render(request, 'main_site/actions/homeworks/delete.html', context)


@login_required(login_url='/accounts/login/')
def homework_delete_page(request, task_id, homework_id):
    task = Task.objects.get(pk=task_id)
    homework = task.homework_set.get(pk=homework_id)
    context = {
        'task_id': task_id,
        'task': task,
        'homework': homework
    }
    return render(request, 'main_site/actions/homeworks/delete_confirm.html', context)


@login_required(login_url='/accounts/login/')
def homework_delete_confirm(request, task_id, homework_id):
    task = Task.objects.get(pk=task_id)
    homework = task.homework_set.get(pk=homework_id)
    try:
        homework.delete()
    except Exception:
        context = {
            'title': "Удаление работы",
            'text': "Произошла ошибка, работы не удалена"
        }
        return render(request, 'main_site/actions/error.html', context)
    else:
        context = {
            'title': "Удаление работы",
            'text': "Работа успешно удалена"
        }
        return render(request, 'main_site/actions/done.html', context)

########################


@login_required(login_url='/accounts/login/')
def task_add(request):
    user_data = UserData.objects.get(user=request.user)
    context = dict()
    context['coggle_key'] = user_data.coggle_key == "Undefined"
    context['miro_key'] = user_data.miro_key == "Undefined"
    return render(request, 'main_site/actions/add.html', context)


@login_required(login_url='/accounts/login/')
def task_add_confirm(request):
    if request.method == 'POST':
        task_form = TaskForm(request.POST, request.FILES)
        if task_form.is_valid():
            file_table = request.FILES['file_table']
            # Загрузка нужной таблицы
            workbook = load_workbook(filename=file_table, read_only=True)
            data = task_form.cleaned_data
            # Проверка наличия заданной таблицы в Excel файле
            if data['table_name'] != '':
                if not checks.check_correct_tablename(workbook, data['table_name']):
                    context = {
                        'title': "Создание задания",
                        'text': "Заданная таблица отсутствует в файле "
                    }
                    return render(request, 'main_site/actions/error.html', context)

            # Проверка формата заданных ячеек
            if data['names_column'] != '' and not checks.column_name_is_correct(data['names_column']) \
                    or data['links_column'] != '' and not checks.column_name_is_correct(data['links_column']) \
                    or data['start_row'] != '' and not checks.column_name_is_correct(data['start_row']):
                context = {
                    'title': "Создание задания",
                    'text': "Ошибка в формате столбцов / ячеек"
                }
                return render(request, 'main_site/actions/error.html', context)

            # Создание массива ключевых значений
            arr_keys = data['text_keys'].split(",")

            if data['table_name'] != '':
                sheet = workbook[data['table_name']]
            else:
                sheet = workbook[workbook.sheetnames[0]]
            print("[SHEET] " + str(sheet))

            # Чтение информации из таблицы
            arr = readers.read_mindmap_ids(
                sheet=sheet,
                names_column=data['names_column'],
                links_column=data['links_column'],
                start_row=data['start_row']
            )
            print("IDs: " + str(arr))

            user_data = UserData.objects.get(user=request.user)
            web_service_object = None
            # Получение массива коэффициентов структурного сходства
            if data['service'] == "Coggle":
                web_service_object = coggle.Coggle(
                    app_name=coggle.APP_NAME,
                    client_id=coggle.CLIENT_ID,
                    client_secret=coggle.CLIENT_SECRET,
                    redirect_uri=coggle.REDIRECT_URI,
                    access_token=user_data.coggle_key
                )
                arr_with_coef = analysis.mindmap_analysis(
                    identificators=arr[1],
                    correct_mindmap_id=readers.link_to_id(data['correct_work']),
                    service=web_service_object
                )
            elif data['service'] == "Miro":
                web_service_object = miro.Miro(
                    app_name=miro.APP_NAME,
                    client_id=miro.CLIENT_ID,
                    client_secret=miro.CLIENT_SECRET,
                    redirect_uri=miro.REDIRECT_URI,
                    access_token=user_data.miro_key
                )
                arr_with_coef = analysis.mindmap_analysis(
                    identificators=arr[1],
                    correct_mindmap_id=readers.link_to_id(data['correct_work']),
                    service=web_service_object
                )
            else:
                context = {
                    'title': "Создание задания",
                    'text': "Ошибка в выборе сервиса"
                }
                return render(request, 'main_site/actions/error.html', context)

            print("Coefs: " + str(arr_with_coef))

            # Массив текстов каждой интеллект карты
            # arr_text = analysis.take_text(arr[1])

            sim_arr = []
            # # Массив с вычисленными сходствами текстовых составляющих
            # if len(arr_keys) > 0:
            #     sim_arr = text_search.initialization(arr_keys, arr_text)
            # else:
            #     sim_arr = [0] * len(arr_keys)

            # Создание объекта домашнего задания для текущего пользователя
            curr_data = UserData.objects.get(user=request.user)
            curr_task = curr_data.task_set.create(title=data['title'], about=data['about'])
            print("[CREATING TASK] TITLE: " + data['title'] + "; ABOUT: " + data['about'])

            # Создание работ
            for i in range(0, len(arr[0])):
                # Проверка на считывание лишней информации
                if arr[0][i] is None:
                    break
                if data['correct_work'] == '':
                    coef = round(arr_with_coef[1][i] * 100)
                else:
                    coef = round(arr_with_coef[i] * 100)

                if len(sim_arr) > 0:
                    sim = round(sim_arr[i] * 100)
                else:
                    sim = -1
                print("[ADDING HW] NAME: " + arr[0][i])
                curr_task.homework_set.create(
                    name=arr[0][i],
                    link=arr[1][i],
                    similarity=coef,
                    plagiarism=sim)
            context = {
                'title': "Создание задания",
                'text': "Задание успешно добавлено"
            }
            return render(request, 'main_site/actions/done.html', context)
        else:
            return render(request, 'main_site/actions/add.html', {'form': task_form})

    else:
        task_form = TaskForm()
    return render(request, 'main_site/actions/add.html', {'form': task_form})


########################

@login_required(login_url='/accounts/login/')
def change(request):
    try:
        user_data = UserData.objects.get(user=request.user)
    except UserData.DoesNotExist:
        user_data = ''
    context = {'userData': user_data}
    return render(request, 'main_site/actions/change.html', context)


@login_required(login_url='/accounts/login/')
def change_task(request, task_id):
    task = Task.objects.get(pk=task_id)
    context = {'task': task,
               'task_id': task_id}
    return render(request, 'main_site/actions/change_confirm.html', context)


@login_required(login_url='/accounts/login/')
def change_task_confirm(request, task_id):
    task = Task.objects.get(pk=task_id)
    try:
        title_new = request.POST['title']
        about_new = request.POST['about']
    except:
        context = {
            'title': "Изменение описания",
            'text': "Произошла ошибка, изменения не сохранены"
        }
        return render(request, 'main_site/actions/error.html', context)
    else:
        task.title = title_new
        task.about = about_new
        task.save()
        context = {
            'title': "Изменение описания",
            'text': "Изменения успешно сохранены"
        }
        return render(request, 'main_site/actions/done.html', context)


########################

@login_required(login_url='/accounts/login/')
def delete(request):
    try:
        user_data = UserData.objects.get(user=request.user)
    except UserData.DoesNotExist:
        user_data = ''
    context = {'userData': user_data}
    return render(request, 'main_site/actions/delete.html', context)


@login_required(login_url='/accounts/login/')
def delete_task(request, task_id):
    task = Task.objects.get(pk=task_id)
    context = {
        'task': task,
        'task_id': task_id
    }
    return render(request, 'main_site/actions/delete_confirm.html', context)


@login_required(login_url='/accounts/login/')
def delete_task_confirm(request, task_id):
    task = Task.objects.get(pk=task_id)
    try:
        task.delete()
    except:
        context = {
            'title': "Удаление задание",
            'text': "Произошла ошибка при удалении задания"
        }
        return render(request, 'main_site/actions/error.html', context)
    else:
        context = {
            'title': "Удаление задания",
            'text': "Задание успешно удалено"
        }
        return render(request, 'main_site/actions/done.html', context)


########################

@login_required(login_url='/accounts/login/')
def coggle_auth_view(request):
    obj = coggle.Coggle(
        app_name=coggle.APP_NAME,
        client_id=coggle.CLIENT_ID,
        client_secret=coggle.CLIENT_SECRET,
        redirect_uri=coggle.REDIRECT_URI
    )
    obj.authorization()
    return redirect(task_add)


@login_required(login_url='/accounts/login/')
def coggle_get_code_view(request):
    try:
        code = request.GET['code']
    except Exception as e:
        print(e)
        context = {
            'title': "Ошибка токена",
            'text': "Ключ авторизации не был получен"
        }
        return render(request, 'main_site/actions/error.html', context)
    else:
        params = {
            'code': code,
            'grant_type': "authorization_code",
            'redirect_uri': coggle.REDIRECT_URI
        }
        response = requests.post(
            coggle.URL_BASE + "token",
            auth=HTTPBasicAuth(coggle.CLIENT_ID, coggle.CLIENT_SECRET),
            json=params
        )
        information_auth = json.loads(response.text)
        user_data = UserData.objects.get(user=request.user)
        user_data.coggle_key = information_auth['access_token']
        user_data.save()
        context = {
            'title': "Получение токена",
            'text': "Ключ авторизации был успешно получен"
        }
        return render(request, 'main_site/actions/done.html', context)


@login_required(login_url='/accounts/login/')
def miro_auth_view(request):
    obj = miro.Miro(
        app_name=miro.APP_NAME,
        client_id=miro.CLIENT_ID,
        client_secret=miro.CLIENT_SECRET,
        redirect_uri=miro.REDIRECT_URI
    )
    obj.authorization()
    return redirect(task_add)


@login_required(login_url='/accounts/login/')
def miro_get_code_view(request):
    try:
        code = request.GET['code']
    except Exception as e:
        print(e)
        context = {
            'title': "Ошибка токена",
            'text': "Ключ авторизации не был получен"
        }
        return render(request, 'main_site/actions/error.html', context)
    else:
        params = {
            'code': code,
            'grant_type': "authorization_code",
            'redirect_uri': miro.REDIRECT_URI
        }
        response = requests.post(
            miro.miro_client.url_base + "token",
            auth=HTTPBasicAuth(miro.CLIENT_ID, miro.CLIENT_SECRET),
            json=params
        )
        information_auth = json.loads(response.text)
        user_data = UserData.objects.get(user=request.user)
        user_data.miro_key = information_auth['access_token']
        user_data.save()
        context = {
            'title': "Получение токена",
            'text': "Ключ авторизации был успешно получен"
        }
        return render(request, 'main_site/actions/done.html', context)
